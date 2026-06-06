"""
schedule_xlsx.py - Generation de fichier Excel template pour l'import de posts planifies.

Specs:
- Colonnes: media_id | date_schedule | feed_visibility | post_action | post_action_delay_seconds | caption
- Sheet "Posts"
- Date format yyyy-mm-dd hh:mm:ss
- post_action="delete", post_action_delay_seconds=172800 (2 jours)
- Heures fixes mais minutes randomisees entre :03 et :25 (jamais sur l'heure pile)
- Media IDs recyclees en ordre
- Captions tirees au hasard
- Filename: template_import_[PRENOM]_[PERIODE].xlsx
"""
from __future__ import annotations

import io
import random
from datetime import datetime, timedelta
from typing import List, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore


# 6 captions par defaut (rotation aleatoire)
DEFAULT_CAPTIONS: List[str] = [
    "Ton abonnement 100% GRATUIT + un CADEAU aujourd'hui seulement \U0001F609❤️",
    "Abonnement gratuit sans code, si tu likes mes derniers posts = \U0001F381",
    "Abonnement gratuit sans code et si tu likes mes 5 derniers posts = cadeau",
    "Abonnement gratuit sans code + surprise si tu likes mes posts \U0001F48B",
    "Abonnement gratuit 0€ + des surprises si tu likes mes derniers post",
    "Abonnement 100% gratuit sans code + des surprises en prive",
]


COLUMNS = [
    "media_id",
    "date_schedule",
    "feed_visibility",
    "post_action",
    "post_action_delay_seconds",
    "caption",
]


def _random_minute() -> int:
    """Retourne une minute aleatoire entre 3 et 25 inclus (jamais sur l'heure)."""
    return random.randint(3, 25)


def _parse_hours(raw: str) -> List[int]:
    """Parse une string "9, 14, 20" -> [9, 14, 20]. Accepte aussi des sauts de ligne."""
    if not raw:
        return []
    out: List[int] = []
    for part in raw.replace("\n", ",").replace(";", ",").split(","):
        part = part.strip()
        if not part:
            continue
        try:
            h = int(part)
            if 0 <= h <= 23:
                out.append(h)
        except ValueError:
            continue
    return out


def _parse_lines(raw: str) -> List[str]:
    """Parse un textarea en liste de lignes non vides."""
    if not raw:
        return []
    return [ln.strip() for ln in raw.splitlines() if ln.strip()]


def build_rows(
    date_start: str,
    date_end: str,
    public_hours: List[int],
    private_hours: List[int],
    media_ids: List[str],
    captions: List[str],
    recycle_infinite: bool = True,
    shuffle_media: bool = False,
    randomize_minutes: bool = True,
) -> List[Tuple[str, str, str, str, int, str]]:
    """Genere la liste des rows pour le xlsx.

    Pour chaque jour de date_start a date_end inclus :
      - 1 post public par heure publique
      - 1 post prive par heure privee
    Les captions sont tirees au hasard.

    Options :
    - recycle_infinite (defaut True) : recycle les medias en boucle quand on
      arrive au bout. Si False : on s arrete des qu on a epuise la liste
      (ex: 3 medias = 3 posts max, peu importe la periode).
    - shuffle_media (defaut False) : melange la liste des medias avant
      utilisation (au lieu de l ordre fourni).
    - randomize_minutes (defaut True) : minutes aleatoires :03 a :25 chaque
      jour pour faire plus humain. Si False : :00 (heure exacte du planning).
    """
    if not media_ids:
        raise ValueError("Aucun media_id fourni")
    if not captions:
        captions = list(DEFAULT_CAPTIONS)

    d_start = datetime.strptime(date_start, "%Y-%m-%d").date()
    d_end = datetime.strptime(date_end, "%Y-%m-%d").date()
    if d_end < d_start:
        raise ValueError("date_end < date_start")

    # Shuffle si demande (copie locale - n affecte pas la liste de l appelant)
    media_ids = list(media_ids)
    if shuffle_media:
        random.shuffle(media_ids)

    rows: List[Tuple[str, str, str, str, int, str]] = []
    media_idx = 0
    total_media = len(media_ids)

    def _minute() -> int:
        return _random_minute() if randomize_minutes else 0

    def _next_media() -> str:
        """Retourne le prochain media. Recycle en boucle si recycle_infinite,
        sinon retourne None quand epuise."""
        nonlocal media_idx
        if recycle_infinite:
            mid = media_ids[media_idx % total_media]
            media_idx += 1
            return mid
        # Pas de recycle : on s arrete quand on est au bout
        if media_idx >= total_media:
            return None
        mid = media_ids[media_idx]
        media_idx += 1
        return mid

    day = d_start
    stop = False
    while day <= d_end and not stop:
        # Public posts
        for h in public_hours:
            mid = _next_media()
            if mid is None:
                stop = True
                break
            m = _minute()
            dt = datetime(day.year, day.month, day.day, h, m, 0)
            rows.append((
                mid,
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                "public",
                "delete",
                172800,
                random.choice(captions),
            ))
        if stop:
            break
        # Private posts
        for h in private_hours:
            mid = _next_media()
            if mid is None:
                stop = True
                break
            m = _minute()
            dt = datetime(day.year, day.month, day.day, h, m, 0)
            rows.append((
                mid,
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                "private",
                "delete",
                172800,
                random.choice(captions),
            ))
        day += timedelta(days=1)

    # Trier par date pour avoir un fichier propre chronologiquement
    rows.sort(key=lambda r: r[1])
    return rows


def generate_xlsx(
    model_name: str,
    date_start: str,
    date_end: str,
    public_hours_raw: str,
    private_hours_raw: str,
    media_ids_raw: str,
    captions_raw: str,
    recycle_infinite: bool = True,
    shuffle_media: bool = False,
    randomize_minutes: bool = True,
) -> Tuple[bytes, str]:
    """Genere le fichier xlsx et retourne (bytes, filename).

    model_name : ex "Amelia"
    date_start, date_end : "YYYY-MM-DD"
    public_hours_raw / private_hours_raw : "9, 14, 20"
    media_ids_raw : un id par ligne
    captions_raw : une caption par ligne (vide -> defauts)
    recycle_infinite / shuffle_media / randomize_minutes : voir build_rows.
    """
    if Workbook is None:
        raise RuntimeError(
            "openpyxl n'est pas installe. Ajoute 'openpyxl>=3.1.0' a requirements.txt "
            "et fais : pip install openpyxl"
        )

    public_hours = _parse_hours(public_hours_raw)
    private_hours = _parse_hours(private_hours_raw)
    media_ids = _parse_lines(media_ids_raw)
    captions = _parse_lines(captions_raw) or list(DEFAULT_CAPTIONS)

    rows = build_rows(
        date_start=date_start,
        date_end=date_end,
        public_hours=public_hours,
        private_hours=private_hours,
        media_ids=media_ids,
        captions=captions,
        recycle_infinite=recycle_infinite,
        shuffle_media=shuffle_media,
        randomize_minutes=randomize_minutes,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Posts"

    # Header
    ws.append(COLUMNS)

    # Data
    for r in rows:
        ws.append(list(r))

    # Largeur de colonnes lisible
    widths = [22, 22, 16, 14, 28, 80]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    prenom = "".join(c for c in (model_name or "model").strip() if c.isalnum() or c in "-_").lower() or "model"
    periode = f"{date_start}_to_{date_end}".replace("-", "")
    filename = f"template_import_{prenom}_{periode}.xlsx"

    return buf.getvalue(), filename
