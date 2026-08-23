# -*- coding: utf-8 -*-
"""Registre des ventes des chatteurs, exportable en Excel.

But : pouvoir verifier une vente contestee. Un chatteur affirme avoir fait
une vente -> on ouvre le fichier, on cherche, et on voit l'heure exacte, la
creatrice concernee, le fan, le montant et le type.

La source est le tableau de transactions de MyPuls (scraping) : c'est le seul
endroit qui donne le detail vente par vente. L'API officielle, elle, ne rend
que des agregats — d'ou l'ecart possible entre les deux, signale dans l'onglet
« Controle ».
"""
from __future__ import annotations

import datetime as _dt
import io
import re
from collections import defaultdict

# Colonnes du registre, dans l'ordre d'affichage
COLONNES = ["Date", "Heure", "Quinzaine", "Chatteur", "Creatrice", "Fan",
            "Montant", "Devise", "Type"]


MOIS_FR = ["", "janvier", "fevrier", "mars", "avril", "mai", "juin", "juillet",
           "aout", "septembre", "octobre", "novembre", "decembre"]


def libelle_quinzaine(cle: str) -> str:
    """« 2026-08 (1-15) » -> « 1 au 15 aout 2026 ».

    La cle reste la forme technique, parce qu'elle se trie toute seule dans
    l'ordre chronologique ; le libelle, lui, est fait pour etre lu.
    """
    t = (cle or "").strip()
    if len(t) < 7:
        return t
    try:
        annee, mois = int(t[:4]), int(t[5:7])
        nom = MOIS_FR[mois]
    except (ValueError, IndexError):
        return t
    if "16" in t:
        import calendar
        fin = calendar.monthrange(annee, mois)[1]
        return "16 au %d %s %d" % (fin, nom, annee)
    return "1 au 15 %s %d" % (nom, annee)


def quinzaine(date_iso: str) -> str:
    """« 2026-08 (1-15) » ou « 2026-08 (16-fin) » — la maille de paie."""
    t = (date_iso or "").strip()
    if len(t) < 10:
        return ""
    try:
        jour = int(t[8:10])
    except ValueError:
        return ""
    return "%s (%s)" % (t[:7], "1-15" if jour <= 15 else "16-fin")


def _quand(brut: str):
    """Coupe la date MyPuls en (date, heure). Formats rencontres :
    « 16/08/2026 14:32 », « 2026-08-16 14:32:05 », ou une date seule."""
    t = (brut or "").strip()
    if not t:
        return "", ""
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d %H:%M"):
        try:
            d = _dt.datetime.strptime(t, fmt)
            return d.strftime("%Y-%m-%d"), d.strftime("%H:%M")
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return _dt.datetime.strptime(t, fmt).strftime("%Y-%m-%d"), ""
        except ValueError:
            pass
    # format inconnu : on garde le brut plutot que de perdre l'information
    m = re.match(r"(\S+)\s+(\S+)", t)
    return (m.group(1), m.group(2)) if m else (t, "")


NON_ATTRIBUE = "(non attribue)"


def _nom_chatteur(brut: str) -> str:
    """Le nom du chatteur, ou NON_ATTRIBUE si MyPuls n'a rattache personne.

    MyPuls ne laisse pas la case vide : il y ecrit « Indetermine (Creatrice) ».
    Pris au pied de la lettre, ce libelle devient un chatteur a part entiere —
    il apparait dans les recaps, se voit calculer une part, et personne ne
    remarque que 20 % du chiffre est en fait orphelin. On le ramene donc a ce
    qu'il est : une absence d'attribution. La creatrice reste lisible dans sa
    propre colonne.
    """
    t = (brut or "").strip()
    if not t:
        return NON_ATTRIBUE
    sans_accent = (t.replace("é", "e").replace("É", "E")
                    .replace("è", "e").replace("ê", "e"))
    if sans_accent.lower().lstrip("( ").startswith("indetermine"):
        return NON_ATTRIBUE
    return t


def est_non_attribue(nom: str) -> bool:
    """Vrai pour toute forme d'absence d'attribution, la notre comme celle de MyPuls."""
    return _nom_chatteur(nom) == NON_ATTRIBUE


def lignes_ventes(transactions) -> list:
    """Transactions MyPuls -> lignes du registre, triees par date decroissante."""
    lignes = []
    for t in (transactions or []):
        d, h = _quand(t.get("date"))
        try:
            montant = round(float(t.get("amount") or 0), 2)
        except (TypeError, ValueError):
            montant = 0.0
        lignes.append([
            d, h, quinzaine(d),
            _nom_chatteur(t.get("chatter")),
            (t.get("creator") or "").strip(),
            (t.get("fan") or "").strip(),
            montant,
            (t.get("currency") or "").strip().upper() or "EUR",
            (t.get("type") or "").strip(),
        ])
    lignes.sort(key=lambda r: (r[0], r[1]), reverse=True)
    return lignes


def lignes_affichables(lignes) -> list:
    """Les memes lignes, quinzaine ecrite en clair.

    Le regroupement et le tri travaillent sur la cle technique, qui se classe
    d'elle-meme dans l'ordre du calendrier ; ce n'est qu'au moment d'ecrire
    dans le classeur qu'on la remplace par « 1 au 15 aout 2026 ».
    """
    return [r[:2] + [libelle_quinzaine(r[2])] + r[3:] for r in lignes]


def recap_par_chatteur(lignes) -> list:
    """Total par chatteur et par devise — la base d'un calcul de paie."""
    par = defaultdict(lambda: defaultdict(float))
    nb = defaultdict(int)
    for r in lignes:
        par[r[3]][r[7]] += r[6]      # [3]=chatteur, [7]=devise, [6]=montant
        nb[r[3]] += 1
    devises = sorted({d for v in par.values() for d in v})
    out = []
    for chatteur in sorted(par, key=lambda c: -sum(par[c].values())):
        ligne = [chatteur, nb[chatteur]]
        ligne += [round(par[chatteur].get(d, 0.0), 2) for d in devises]
        out.append(ligne)
    return ["Chatteur", "Ventes"] + ["Total " + d for d in devises], out


def recap_par_quinzaine(lignes) -> tuple:
    """Total par chatteur ET par quinzaine : c'est ce qui sert a payer."""
    par = defaultdict(lambda: defaultdict(float))
    for r in lignes:
        if r[2]:
            par[r[3]][r[2]] += r[6]
    quinz = sorted({q for v in par.values() for q in v})
    out = []
    for chatteur in sorted(par, key=lambda c: -sum(par[c].values())):
        out.append([chatteur] + [round(par[chatteur].get(q, 0.0), 2) for q in quinz]
                   + [round(sum(par[chatteur].values()), 2)])
    # Le tri s'est fait sur la cle technique (chronologique) ; l'en-tete, lui,
    # s'affiche en clair.
    return (["Chatteur"] + [libelle_quinzaine(q) for q in quinz] + ["Total"]), out


def fiches_chatteurs(lignes) -> list:
    """Une fiche par chatteur : ses chiffres, puis le detail de ses ventes.

    C'est la vue qu'on ouvre quand quelqu'un conteste : on voit d'un coup ce
    qu'il a fait, sa part du chiffre, son panier moyen, puis chaque vente avec
    sa date, son heure, le compte sur lequel elle a ete faite et le fan.
    """
    par = defaultdict(list)
    for r in lignes:
        par[r[3]].append(r)
    total_general = sum(r[6] for r in lignes) or 1.0
    out = []
    for chatteur in sorted(par, key=lambda c: -sum(r[6] for r in par[c])):
        ventes = par[chatteur]
        total = sum(r[6] for r in ventes)
        panier = total / len(ventes) if ventes else 0.0
        part = 100.0 * total / total_general
        # en-tete de la fiche
        out.append([chatteur, "%d vente(s)" % len(ventes), round(total, 2),
                    "%.1f %% du total" % part, "panier %.2f" % panier])
        # ses ventes, de la plus recente a la plus ancienne
        for r in sorted(ventes, key=lambda x: (x[0], x[1]), reverse=True):
            out.append(["", r[0] + " " + r[1], r[4], r[5], "%.2f %s" % (r[6], r[7])])
        out.append(["", "", "", "", ""])          # respiration entre deux fiches
    return out


def _categorie_vente(type_vente: str) -> str:
    """Famille d'une vente, decidee par mypuls et par lui seul.

    La regle etait recopiee ici : la copie cherchait « ppv » alors que MyPuls
    ecrit « Media prive », donc la colonne PPV affichait zero et PPV + Tips ne
    recoupaient jamais le CA. Import tardif pour ne pas charger le scraper
    quand on ne fait que relire un registre.
    """
    from mypuls import categorie_transaction
    return categorie_transaction(type_vente)


# Colonnes fixes de la feuille de paie, avant les colonnes « Vente N ».
COLONNES_PAIE = ["Quinzaine", "Chatteur", "Ventes", "CA total (USD)",
                 "PPV (USD)", "Tips (USD)", "Autres (USD)", "CA EUR", "CA USD",
                 "Commission %", "A payer (USD)"]


def paie_quinzaine(lignes, commissions=None, eur_usd: float = 1.14) -> tuple:
    """Une ligne par chatteur et par quinzaine : CA, taux, a payer, et le
    detail de toutes ses ventes SUR LA MEME LIGNE.

    `commissions` : {chatteur: pourcentage}. La part EUR est convertie en USD,
    la part USD est prise telle quelle — meme regle que la page Paie du site,
    sinon les ventes OnlyFans seraient surpayees.

    TOUS les montants de synthese sont en dollars : additionner des euros MyM
    et des dollars OnlyFans dans une colonne « CA total » gonflait la part
    OnlyFans d'environ 6,5 %. Les colonnes « CA EUR » et « CA USD » gardent, elles,
    les montants bruts dans leur devise d'origine — c'est la matiere premiere.

    « PPV » = le contenu payant (messages prives + publications), « Tips » les
    pourboires, « Autres » TOUT le reste (abonnements, streams, parrainages, et
    les libelles qu'aucune famille ne reconnait). Les trois colonnes somment
    exactement le CA total : une vente ne peut plus disparaitre entre elles.
    """
    commissions = commissions or {}

    def _en_usd(v):
        """Un montant de vente ramene en dollars. USD = OnlyFans, tout le reste
        est traite comme du MyM (meme repli que mypuls._norm_currency)."""
        return v[6] * (1.0 if v[7] == "USD" else eur_usd)

    par = defaultdict(list)
    for r in lignes:
        if r[2]:
            par[(r[3], r[2])].append(r)
    out = []
    for (chatteur, quinz) in sorted(par, key=lambda k: (k[1], -sum(x[6] for x in par[k]))):
        ventes = par[(chatteur, quinz)]
        ca_eur = sum(v[6] for v in ventes if v[7] == "EUR")
        ca_usd = sum(v[6] for v in ventes if v[7] == "USD")
        autres_devises = sum(v[6] for v in ventes if v[7] not in ("EUR", "USD"))
        ca = sum(_en_usd(v) for v in ventes)
        # Meme decoupage que les cartes du site, meme regle de classement.
        # La famille est calculee UNE fois par vente : elle sert deux colonnes.
        familles = [(_categorie_vente(v[8]), _en_usd(v)) for v in ventes]
        ppv = sum(m for f, m in familles if f in ("Messages", "Posts"))
        tips = sum(m for f, m in familles if f == "Tips")
        # Le reste par SOUSTRACTION : ce qu'aucune famille ne reconnait reste
        # visible dans la ligne au lieu de manquer au total.
        autres_types = ca - ppv - tips
        pct = float(commissions.get(chatteur, commissions.get("_defaut_", 0)) or 0)
        a_payer = round(ca * pct / 100.0, 2)
        # Les ventes s'etalent vers la droite, UNE PAR CASE, de la plus
        # recente a la plus ancienne. Entassees dans une seule cellule elles
        # etaient illisibles et impossibles a trier ou filtrer.
        detail = ["%s %.2f%s (%s)" % (v[0][5:], v[6],
                                      "€" if v[7] == "EUR" else "$", v[4])
                  for v in sorted(ventes, key=lambda x: (x[0], x[1]), reverse=True)]
        out.append([libelle_quinzaine(quinz), chatteur, len(ventes),
                    round(ca, 2), round(ppv, 2), round(tips, 2),
                    round(autres_types, 2),
                    # Une devise exotique est comptee avec les euros, comme
                    # partout ailleurs (mypuls._norm_currency retombe sur EUR) :
                    # ainsi CA EUR + CA USD reste la matiere du CA total.
                    round(ca_eur + autres_devises, 2), round(ca_usd, 2),
                    pct, a_payer] + detail)

    # Autant de colonnes « Vente N » que le chatteur le plus actif en a.
    fixes = len(COLONNES_PAIE)
    largeur = max((len(r) - fixes for r in out), default=0)
    cols = COLONNES_PAIE + ["Vente %d" % (i + 1) for i in range(largeur)]
    return cols, out


def rapprochement(lignes, chatters=None, diagnostic=None) -> tuple:
    """Confronte le tableau du site au registre, chatteur par chatteur.

    Deux sources cohabitent chez MyPuls : la table de performance (celle
    qu'affiche la page Revenus) et le log de transactions (celui qui detaille
    chaque vente). Elles ne disent pas toujours la meme chose — un chatteur
    peut peser sur la premiere et n'avoir presque rien dans la seconde.

    Tant que personne ne les comparait, l'ecart restait invisible : c'est
    exactement le « des ventes qui ne montent pas ». Ici on met les deux
    colonnes cote a cote et on nomme le coupable.
    """
    cols = ["Chatteur", "CA affiche sur le site", "Ventes au registre",
            "Total du registre", "Ecart", "Verdict"]
    # Le CA de la table perf additionne EUR et USD dans la meme colonne : on
    # compare donc a la somme BRUTE du registre, sans conversion.
    par_nom, nb = defaultdict(float), defaultdict(int)
    for r in lignes:
        cle = (r[3] or "").strip().lower()
        par_nom[cle] += r[6]
        nb[cle] += 1

    out, vus = [], set()
    for c in (chatters or []):
        nom = (c.get("name") or "").strip()
        # Cote site aussi, « Indetermine (X) » se fait passer pour un chatteur :
        # on le ramene au meme libellé que dans le registre, sinon les deux
        # colonnes ne parlent pas de la meme chose.
        cle = _nom_chatteur(nom).lower()
        if cle == NON_ATTRIBUE:
            nom = "%s — %s" % (NON_ATTRIBUE, nom)
        vus.add(cle)
        site = float(c.get("ca_total") or 0)
        reg = round(par_nom.get(cle, 0.0), 2)
        ecart = round(reg - site, 2)
        # 5 % de tolerance : les deux sources n'arrondissent pas pareil.
        if cle == NON_ATTRIBUE:
            verdict = "NON ATTRIBUE par MyPuls — ne pas payer"
            out.append([nom, round(site, 2), nb.get(cle, 0), reg, ecart, verdict])
            continue
        if not par_nom.get(cle) and site > 0:
            verdict = "ABSENT du registre — aucune vente detaillee"
        elif abs(ecart) <= max(1.0, 0.05 * site):
            verdict = "coherent"
        elif ecart < 0:
            verdict = "INCOMPLET — il manque des ventes au registre"
        else:
            verdict = "registre plus fourni que le site"
        out.append([nom, round(site, 2), nb.get(cle, 0), reg, ecart, verdict])

    # Presents dans le registre, absents du tableau du site : ceux-la sont
    # payes nulle part si on ne regarde que la page Revenus.
    for cle, tot in sorted(par_nom.items(), key=lambda kv: -kv[1]):
        if cle in vus:
            continue
        anonyme = (not cle) or cle.startswith("(")
        nom = cle if anonyme else cle.title()
        out.append([nom or "(non attribue)", "", nb[cle], round(tot, 2), "",
                    "ventes sans nom — a rattacher"
                    if anonyme else "ABSENT du tableau du site"])

    out.sort(key=lambda r: (r[5] == "coherent", -(r[3] or 0)))

    if diagnostic:
        out.append(["", "", "", "", "", ""])
        out.append(["Lecture du log MyPuls", "", "", "", "", ""])
        etiquettes = [
            ("ventes_lues", "Ventes lues dans le log"),
            ("lignes_illisibles", "Lignes du log NON lues (colonnes manquantes)"),
            ("chatters_illisibles", "Lignes du tableau chatteurs non lues"),
            ("ventes_sans_chatteur", "Ventes sans chatteur"),
            ("montant_sans_chatteur", "Montant sans chatteur"),
        ]
        for cle, libelle in etiquettes:
            if cle in diagnostic:
                out.append([libelle, diagnostic[cle], "", "", "", ""])
    return cols, out


def controle(lignes, total_annonce=None) -> list:
    """Compare la somme du registre au total affiche ailleurs.

    C'est ce qui repond a « des fois il y a des trucs qui ne montent pas » :
    si l'ecart n'est pas nul, une vente manque quelque part.
    """
    par_devise = defaultdict(float)
    sans_chatteur = sans_date = 0
    for r in lignes:
        par_devise[r[7]] += r[6]
        if r[3] == "(non attribue)":
            sans_chatteur += 1
        if not r[0]:
            sans_date += 1
    # --- recherche de trous : un jour sans AUCUNE vente au milieu d'une
    # periode active est suspect, c'est la trace d'un scraping incomplet.
    jours = defaultdict(float)
    for r in lignes:
        if r[0]:
            jours[r[0]] += r[6]
    trous = []
    if len(jours) >= 3:
        tous = sorted(jours)
        try:
            d0 = _dt.date.fromisoformat(tous[0])
            d1 = _dt.date.fromisoformat(tous[-1])
            j = d0
            while j <= d1:
                if j.isoformat() not in jours:
                    trous.append(j.isoformat())
                j += _dt.timedelta(days=1)
        except Exception:
            pass

    ctrl = [["Ventes enregistrees", len(lignes)]]
    for d, v in sorted(par_devise.items()):
        ctrl.append(["Total " + d, round(v, 2)])
    ctrl.append(["Ventes sans chatteur identifie", sans_chatteur])
    ctrl.append(["Ventes sans date exploitable", sans_date])
    ctrl.append(["Jours couverts", len(jours)])
    ctrl.append(["Jours SANS aucune vente", len(trous)])
    if trous:
        ctrl.append(["", ""])
        ctrl.append(["Jours vides — a verifier sur MyPuls", ""])
        for j in trous[:40]:
            ctrl.append(["", j])
        if len(trous) > 40:
            ctrl.append(["", "… et %d autre(s)" % (len(trous) - 40)])
    # journee la plus faible : utile pour reperer un scraping tronque
    if jours:
        mini = min(jours, key=lambda k: jours[k])
        maxi = max(jours, key=lambda k: jours[k])
        ctrl.append(["", ""])
        ctrl.append(["Journee la plus faible", "%s (%.2f)" % (mini, jours[mini])])
        ctrl.append(["Journee la plus forte", "%s (%.2f)" % (maxi, jours[maxi])])
    if total_annonce is not None:
        somme = round(sum(par_devise.values()), 2)
        ctrl.append(["Total annonce par le dashboard", round(total_annonce, 2)])
        ctrl.append(["Ecart", round(somme - total_annonce, 2)])
    return ctrl


def construire_xlsx(transactions, periode: str = "", total_annonce=None,
                    commissions=None, eur_usd: float = 1.14,
                    chatters=None, diagnostic=None) -> bytes:
    """Le classeur complet.

    `chatters` est le tableau de performance affiche par le site : le fournir
    active la feuille « Site vs registre », qui dit quel chatteur manque de
    quel cote. Sans lui, le classeur reste valable, simplement muet sur ce
    point.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    lignes = lignes_ventes(transactions)
    wb = Workbook()

    gras = Font(bold=True, color="FFFFFF")
    fond = PatternFill("solid", fgColor="1C1C1E")
    centre = Alignment(horizontal="center")

    def _entete(ws, cols):
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            c = ws.cell(row=1, column=i)
            c.font, c.fill, c.alignment = gras, fond, centre
        ws.freeze_panes = "A2"

    # --- feuille 1 : le registre ---
    ws = wb.active
    ws.title = "Ventes"
    _entete(ws, COLONNES)
    for r in lignes_affichables(lignes):
        ws.append(r)
    if lignes:
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(COLONNES)), len(lignes) + 1)
    for i, c in enumerate(COLONNES, start=1):
        largeur = {"Date": 12, "Heure": 8, "Quinzaine": 16, "Chatteur": 22,
                   "Creatrice": 20, "Fan": 24, "Montant": 12, "Devise": 9,
                   "Type": 16}.get(c, 14)
        ws.column_dimensions[get_column_letter(i)].width = largeur
    for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
        for c in row:
            c.number_format = "#,##0.00"

    # --- feuille 2 : recap par chatteur ---
    ws2 = wb.create_sheet("Par chatteur")
    cols2, recap = recap_par_chatteur(lignes)
    _entete(ws2, cols2)
    for r in recap:
        ws2.append(r)
    ws2.column_dimensions["A"].width = 22
    for i in range(2, len(cols2) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 14
    for row in ws2.iter_rows(min_row=2, min_col=3):
        for c in row:
            c.number_format = "#,##0.00"

    # --- feuille 3 : par quinzaine (la maille de paie) ---
    ws4 = wb.create_sheet("Par quinzaine")
    cols4, recap4 = recap_par_quinzaine(lignes)
    _entete(ws4, cols4)
    for r in recap4:
        ws4.append(r)
    ws4.column_dimensions["A"].width = 22
    for i in range(2, len(cols4) + 1):
        ws4.column_dimensions[get_column_letter(i)].width = 16
    for row in ws4.iter_rows(min_row=2, min_col=2):
        for c in row:
            c.number_format = "#,##0.00"

    # --- feuille : la paie, une ligne par chatteur et par quinzaine ---
    ws6 = wb.create_sheet("Paie quinzaine")
    cols6, paie = paie_quinzaine(lignes, commissions=commissions, eur_usd=eur_usd)
    _entete(ws6, cols6)
    for r in paie:
        ws6.append(r)
    for col, larg in zip("ABCDEFGHIJK", (20, 24, 8, 13, 11, 11, 12, 11, 11, 13, 15)):
        ws6.column_dimensions[col].width = larg
    # les colonnes « Vente N » : assez larges pour « 08-15 114.96€ (Emmabrn) »
    _fixes = len(COLONNES_PAIE)
    for i in range(_fixes + 1, len(cols6) + 1):
        ws6.column_dimensions[get_column_letter(i)].width = 26
    for row in ws6.iter_rows(min_row=2, min_col=4, max_col=_fixes):
        for c in row:
            c.number_format = "#,##0.00"
    ws6.freeze_panes = "C2"          # nom et quinzaine restent visibles

    # --- feuille : une fiche par chatteur ---
    ws5 = wb.create_sheet("Fiches chatteurs")
    _entete(ws5, ["Chatteur", "Quand", "Compte", "Fan", "Montant"])
    from openpyxl.styles import Font as _F
    for r in fiches_chatteurs(lignes):
        ws5.append(r)
        if r[0]:                                   # ligne d'en-tete de fiche
            for c in ws5[ws5.max_row]:
                c.font = _F(bold=True)
    for col, larg in zip("ABCDE", (24, 20, 20, 24, 18)):
        ws5.column_dimensions[col].width = larg

    # --- feuille : le site face au registre ---
    # Placee AVANT le controle : c'est la premiere chose a regarder quand on
    # soupconne des ventes manquantes, puisqu'elle nomme les chatteurs
    # concernes au lieu de donner un ecart global.
    ws7 = wb.create_sheet("Site vs registre")
    cols7, rappro = rapprochement(lignes, chatters, diagnostic)
    _entete(ws7, cols7)
    from openpyxl.styles import PatternFill as _PF, Font as _Fo
    alerte = _PF("solid", fgColor="FDE2E1")
    for r in rappro:
        ws7.append(r)
        verdict = str(r[5] or "")
        if verdict.startswith(("ABSENT", "INCOMPLET")):
            for c in ws7[ws7.max_row]:
                c.fill = alerte
        elif not r[1] and not r[2] and r[0]:       # intitules de bas de feuille
            ws7.cell(row=ws7.max_row, column=1).font = _Fo(bold=True)
    for col, larg in zip("ABCDEF", (26, 22, 18, 18, 12, 42)):
        ws7.column_dimensions[col].width = larg
    for row in ws7.iter_rows(min_row=2, min_col=2, max_col=5):
        for c in row:
            c.number_format = "#,##0.00"

    # --- feuille 5 : controle ---
    ws3 = wb.create_sheet("Controle")
    _entete(ws3, ["Verification", "Valeur"])
    if periode:
        ws3.append(["Periode", periode])
    for r in controle(lignes, total_annonce):
        ws3.append(r)
    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
